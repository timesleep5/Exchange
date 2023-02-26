from typing import List

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class Workbookutils:
    @staticmethod
    def fill_row(ws, startcol: int, endcol: int, row: int, content: List[str]):
        for i in range(startcol, endcol + 1):
            char = get_column_letter(i)
            ws[f'{char}{row}'] = content[i - startcol]  # to be at Index 0 in the first iteration

    @staticmethod
    def fill_col(ws, startrow: int, endrow: int, col: int, content: List[str]):
        char = get_column_letter(col)
        for i in range(startrow, endrow + 1):
            ws[f'{char}{i}'] = content[i - startrow]

    @staticmethod
    def make_row_bold(ws, startcol, endcol, row):
        for i in range(startcol, endcol + 1):
            char = get_column_letter(i)
            ws[f'{char}{row}'].font = Font(bold=True)

    @staticmethod
    def make_col_bold(ws, startrow, endrow, col):
        for i in range(startrow, endrow + 1):
            ws[f'{col}{i}'].font = Font(bold=True)
