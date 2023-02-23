from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class CreateWorkbook:
    def __init__(self):
        self.wb = self._get_wb()
        self.week = ['Monday', 'Tuesday', 'Wednesday', 'Tuesday', 'Friday', 'Saturday', 'Sunday']

    @staticmethod
    def _get_wb():
        return Workbook()

    def get_workheets(self):
        self.wb.create_sheet()
        self.wb.create_sheet()

    def style_backlog(self):
        ws = self.wb['Sheet']
        ws.title = 'Backlog'
        ws.append(self.week)
        self._make_row_bold(ws, 1, len(self.week), 1)

    def style_weektable(self):
        ws = self.wb['Sheet1']
        ws.title = 'WeekTable'
        ws.append(self.week)
        self._make_row_bold(ws, 1, len(self.week), 1)

    def _make_row_bold(self, ws, startcol, endcol, row):
        for i in range(startcol, endcol):
            char = get_column_letter(i)
            ws[f'{char}{row}'].font = Font(bold=True)

    def savewb(self):
        self.wb.save(filename='FlowOpt.xlsx')


wb = CreateWorkbook()
wb.get_workheets()
wb.style_backlog()
wb.style_weektable()
wb.savewb()




