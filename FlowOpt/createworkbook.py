from openpyxl import Workbook

from workbookutils import Workbookutils as utils


class CreateWorkbook:
    def __init__(self, wbname):
        self.wb = self._create_wb()
        self.week = ['Monday', 'Tuesday', 'Wednesday', 'Tuesday', 'Friday', 'Saturday', 'Sunday']
        self.categories = ['Day', 'Date', 'Planned', 'Done', 'Resumé']
        self.wbname = wbname
        self.create()

    @staticmethod
    def _create_wb():
        return Workbook()

    def get_workheets(self, amount):
        for _ in range(amount):
            self.wb.create_sheet()

    def style_sheets(self):
        self.style_backlog()
        self.style_schedule()
        self.style_history()

    def style_backlog(self):
        ws = self.wb['Sheet']
        ws.title = 'Backlog'
        ws.append(self.week)
        utils.make_row_bold(ws, 1, len(self.week), 1)

    def style_schedule(self):
        ws = self.wb['Sheet1']
        ws.title = 'Schedule'
        ws.append(self.week)
        utils.make_row_bold(ws, 1, len(self.week), 1)

    def style_history(self):
        ws = self.wb['Sheet2']
        ws.title = 'History'
        for i in range(1, len(self.categories) + 1):
            ws[f'A{i}'] = self.categories[i - 1]
        utils.fill_row(ws, 2, len(self.week) + 1, 1, self.week)
        utils.fill_col(ws, 1, len(self.categories), 1, self.categories)
        utils.make_col_bold(ws, 1, len(self.categories), 'A')

    def create(self):
        self.get_workheets(2)
        self.style_sheets()
        self.savewb()

    def savewb(self):
        self.wb.save(filename=self.wbname)

    def get_workbook(self):
        return self.wbname
